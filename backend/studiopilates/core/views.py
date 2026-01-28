from datetime import date, datetime
import calendar
import json
from io import BytesIO
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth import get_user_model
from django.utils.text import slugify
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.http import HttpResponse
from datetime import timedelta

from . import forms, models, services
from .signals import ensure_profissional_for_user
from shared.ai.gemini_client import extract_address_from_proof, extract_student_from_document


def _active_menu(path: str) -> str:
    if path.startswith("/agenda"):
        return "agenda"
    if path.startswith("/financeiro"):
        return "financeiro"
    if path.startswith("/configuracoes"):
        return "configuracoes"
    if path.startswith("/contratos") or path.startswith("/wizard"):
        return "cadastros"
    if path.startswith("/cadastros"):
        return "cadastros"
    return "dashboard"


def login_view(request):
    if request.method == "POST":
        user = authenticate(request, username=request.POST.get("username"), password=request.POST.get("password"))
        if user:
            ensure_profissional_for_user(user)
            login(request, user)
            return redirect("dashboard")
        messages.error(request, "Login invalido")
    return render(request, "login.html")


def logout_view(request):
    logout(request)
    return redirect("login")


@login_required
def perfil_view(request):
    profissional = models.Profissional.objects.filter(user=request.user).first()
    if not profissional:
        profissional = ensure_profissional_for_user(request.user)
    if not profissional:
        messages.error(request, "Nao foi possivel localizar o seu cadastro.")
        return redirect("dashboard")
    return edit_view(request, models.Profissional, forms.ProfissionalForm, "perfil", profissional.pk)


@login_required
def dashboard(request):
    context = {
        "alunos": models.Aluno.objects.count(),
        "contratos": models.Contrato.objects.count(),
        "reservas_hoje": models.Reserva.objects.filter(aulaSessao__data=date.today()).count(),
        "receber_aberto": models.ContasReceber.objects.filter(status="ABERTO").count(),
        "breadcrumbs": [("Home", "#")],
        "active_menu": "dashboard",
    }
    return render(request, "dashboard.html", context)

@login_required
def aluno_detail(request, pk):
    aluno = get_object_or_404(models.Aluno, pk=pk)
    endereco = aluno.cdEndereco
    telefones = list(aluno.telefones.values_list("dsTelefone", flat=True))
    contratos = models.Contrato.objects.filter(cdAluno=aluno).select_related("cdPlano", "cdUnidade")
    contrato_forms = {contrato.id: forms.ContratoForm(instance=contrato) for contrato in contratos}
    reservas = (
        models.Reserva.objects.filter(aluno=aluno)
        .select_related("aulaSessao", "aulaSessao__profissional", "aulaSessao__unidade", "aulaSessao__tipoServico")
        .order_by("aulaSessao__data", "aulaSessao__horaInicio")
    )
    reserva_forms = {reserva.id: forms.ReservaForm(instance=reserva) for reserva in reservas}
    evolucoes = (
        models.EvolucaoAluno.objects.filter(reserva__aluno=aluno)
        .select_related("profissional", "reserva")
        .order_by("-dtEvolucao")
    )
    contas_receber = _filtrar_contas_receber(
        models.ContasReceber.objects.filter(contrato__cdAluno=aluno).select_related("contrato"),
        request,
    )
    planos = models.Plano.objects.select_related("cdTipoServico").all()
    unidades = models.Unidade.objects.all()
    profissionais = models.Profissional.objects.all()
    context = {
        "aluno": aluno,
        "endereco": endereco,
        "telefones": telefones,
        "contratos": contratos,
        "contrato_forms": contrato_forms,
        "reservas": reservas,
        "reserva_forms": reserva_forms,
        "evolucoes": evolucoes,
        "contas_receber": contas_receber,
        "filtros_financeiro": _get_filtros_financeiro(request),
        "today": timezone.now().date().strftime("%Y-%m-%d"),
        "planos": planos,
        "unidades": unidades,
        "profissionais": profissionais,
        "edit_form": forms.AlunoForm(instance=aluno),
        "breadcrumbs": [("Home", reverse("dashboard")), ("Alunos", reverse("alunos_list")), ("Ficha", "#")],
        "active_menu": "cadastros",
    }
    return render(request, "alunos/detail.html", context)


def _sync_user_for_profissional(profissional, raw_password=None, old_cd=None):
    User = get_user_model()
    base_username = slugify(profissional.profissional) or f"user-{profissional.id}"
    username = base_username
    counter = 1
    while User.objects.filter(username=username).exclude(pk=getattr(profissional.user, "pk", None)).exists():
        counter += 1
        username = f"{base_username}-{counter}"

    if profissional.user_id:
        user = profissional.user
        user.username = username
        user.first_name = profissional.profissional
    else:
        user = User(username=username, first_name=profissional.profissional)

    if raw_password:
        user.set_password(raw_password)
    elif not user.has_usable_password():
        user.set_unusable_password()
    user.save()

    if not profissional.user_id:
        profissional.user = user
        profissional.save(update_fields=["user"])


def _inject_cd_value(model, data):
    cd_field = next((f.name for f in model._meta.fields if f.name.startswith("cd")), None)
    if not cd_field:
        return data
    if data.get(cd_field):
        return data
    max_val = model.objects.order_by(f"-{cd_field}").values_list(cd_field, flat=True).first() or 0
    data[cd_field] = max_val + 1
    return data


def _sync_aluno_address(aluno, data):
    fields = {
        "dsLogradouro": data.get("dsLogradouro", "").strip(),
        "dsNumero": data.get("dsNumero", "").strip(),
        "dsCEP": data.get("dsCEP", "").strip(),
        "dsCidade": data.get("dsCidade", "").strip(),
        "dsBairro": data.get("dsBairro", "").strip(),
    }
    if not any(fields.values()):
        return
    endereco = aluno.cdEndereco
    if not endereco:
        max_cd = models.EnderecoAluno.objects.order_by("-cdEndereco").values_list("cdEndereco", flat=True).first() or 0
        endereco = models.EnderecoAluno(cdEndereco=max_cd + 1, cdAluno=aluno)
    for key, value in fields.items():
        setattr(endereco, key, value)
    endereco.save()
    if aluno.cdEndereco_id != endereco.id:
        aluno.cdEndereco = endereco
        aluno.save(update_fields=["cdEndereco"])


def _sync_aluno_phones(aluno, data):
    phones = []
    for key, value in data.items():
        if key.startswith("telefone_"):
            raw = value.strip()
            if raw:
                phones.append(raw)
    if not phones:
        return
    models.TelefoneAluno.objects.filter(cdAluno=aluno).delete()
    max_cd = models.TelefoneAluno.objects.order_by("-cdTelefone").values_list("cdTelefone", flat=True).first() or 0
    for idx, numero in enumerate(phones, start=1):
        models.TelefoneAluno.objects.create(cdTelefone=max_cd + idx, cdAluno=aluno, dsTelefone=numero)


def _to_time(raw_value):
    if not raw_value:
        return None
    if isinstance(raw_value, str):
        try:
            return datetime.strptime(raw_value.strip(), "%H:%M").time()
        except ValueError:
            return None
    return raw_value


def _add_months(base_date, months):
    year = base_date.year + (base_date.month - 1 + months) // 12
    month = (base_date.month - 1 + months) % 12 + 1
    day = min(base_date.day, 28)
    return date(year, month, day)


def _add_years(base_date, years):
    return _add_months(base_date, years * 12)


def _first_last_day_month(ref_date):
    first = ref_date.replace(day=1)
    last_day = calendar.monthrange(ref_date.year, ref_date.month)[1]
    last = ref_date.replace(day=last_day)
    return first, last


@login_required
def gerar_horarios_studio(request):
    if request.method != "POST":
        return redirect("horarios_studio_list")
    unidade_id = request.POST.get("unidade") or ""
    tipo_id = request.POST.get("tipoServico") or ""
    profissional_id = request.POST.get("profissional") or ""
    dias = request.POST.getlist("dias")
    inicio = _to_time(request.POST.get("horaInicio"))
    fim = _to_time(request.POST.get("horaFim"))
    intervalo = int(request.POST.get("intervalo") or 0)
    capacidade = request.POST.get("capacidade") or ""
    try:
        capacidade = int(capacidade) if capacidade else None
    except ValueError:
        capacidade = None
    if not (unidade_id and tipo_id and dias and inicio and fim and intervalo):
        messages.error(request, "Preencha unidade, tipo de servico, dias e horarios.")
        return redirect("horarios_studio_list")
    try:
        unidade_id = int(unidade_id)
        tipo_id = int(tipo_id)
    except ValueError:
        messages.error(request, "Unidade ou tipo de servico invalido.")
        return redirect("horarios_studio_list")
    prof_id = None
    if profissional_id:
        try:
            prof_id = int(profissional_id)
        except ValueError:
            prof_id = None
    start_minutes = inicio.hour * 60 + inicio.minute
    end_minutes = fim.hour * 60 + fim.minute
    if end_minutes <= start_minutes:
        messages.error(request, "Horario final deve ser maior que o inicial.")
        return redirect("horarios_studio_list")
    if intervalo <= 0:
        messages.error(request, "Intervalo invalido.")
        return redirect("horarios_studio_list")
    max_cd = models.HorarioStudio.objects.order_by("-cdHorario").values_list("cdHorario", flat=True).first() or 0
    created = 0
    for dia in dias:
        try:
            dia_int = int(dia)
        except ValueError:
            continue
        current = start_minutes
        while current + intervalo <= end_minutes:
            hora_inicio = datetime.strptime(f"{current // 60:02d}:{current % 60:02d}", "%H:%M").time()
            next_min = current + intervalo
            hora_fim = datetime.strptime(f"{next_min // 60:02d}:{next_min % 60:02d}", "%H:%M").time()
            exists = models.HorarioStudio.objects.filter(
                unidade_id=unidade_id,
                tipoServico_id=tipo_id,
                profissional_id=prof_id,
                diaSemana=dia_int,
                horaInicio=hora_inicio,
                horaFim=hora_fim,
            ).exists()
            if not exists:
                max_cd += 1
                models.HorarioStudio.objects.create(
                    cdHorario=max_cd,
                    unidade_id=unidade_id,
                    tipoServico_id=tipo_id,
                    profissional_id=prof_id,
                    diaSemana=dia_int,
                    horaInicio=hora_inicio,
                    horaFim=hora_fim,
                    capacidade=capacidade,
                )
                created += 1
            current = next_min
    if created:
        messages.success(request, f"{created} horarios criados.")
    else:
        messages.info(request, "Nenhum horario novo criado.")
    return redirect("horarios_studio_list")


def list_view(request, model, form_class, title, allow_modal=True, extra_context=None):
    query = request.GET.get("q", "").strip()
    order = request.GET.get("order", "id")
    qs = model.objects.all()
    if query:
        field_name = model._meta.fields[1].name
        qs = qs.filter(Q(**{f"{field_name}__icontains": query}) | Q(id__icontains=query))
    if order:
        qs = qs.order_by(order)
    paginator = Paginator(qs, 10)
    page = paginator.get_page(request.GET.get("page"))
    display_fields = [
        {"name": field.name, "label": str(field.verbose_name)}
        for field in model._meta.fields
        if not field.primary_key and not field.name.startswith("cd")
    ][:3]
    edit_forms = {}
    if allow_modal:
        for obj in page:
            edit_forms[obj.id] = form_class(instance=obj)

    context = {
        "title": title,
        "page": page,
        "form": form_class(),
        "model_name": model._meta.model_name,
        "breadcrumbs": [("Home", reverse("dashboard")), (title, "#")],
        "active_menu": _active_menu(request.path),
        "allow_modal": allow_modal,
        "display_fields": display_fields,
        "edit_forms": edit_forms,
    }
    if model is models.Aluno:
        address_map = {}
        phones_map = {}
        for obj in page:
            address_map[obj.id] = obj.cdEndereco
            phones_map[obj.id] = list(obj.telefones.values_list("dsTelefone", flat=True))
        context.update({"address_map": address_map, "phones_map": phones_map})
    if extra_context:
        context.update(extra_context)
    return render(request, "generic/list.html", context)


@login_required
def horarios_studio_list(request):
    extra_context = {
        "unidades": models.Unidade.objects.all(),
        "tipos_servico": models.TipoServico.objects.all(),
        "profissionais": models.Profissional.objects.all(),
        "dias_semana": models.HorarioStudio.DIAS_SEMANA,
    }
    return list_view(
        request,
        models.HorarioStudio,
        forms.HorarioStudioForm,
        "Horario do Studio",
        extra_context=extra_context,
    )


@login_required
def contas_pagar_list(request):
    today = timezone.now().date()
    inicio = request.GET.get("inicio", "").strip()
    fim = request.GET.get("fim", "").strip()
    status = request.GET.get("status", "").strip()
    fornecedor_id = request.GET.get("fornecedor", "").strip()
    categoria_id = request.GET.get("categoria", "").strip()
    subcategoria_id = request.GET.get("subcategoria", "").strip()
    if not inicio or not fim:
        first, last = _first_last_day_month(today)
        if not inicio:
            inicio = first.strftime("%Y-%m-%d")
        if not fim:
            fim = last.strftime("%Y-%m-%d")

    qs = models.ContasPagar.objects.select_related("cdFornecedor", "cdCategoria", "cdSubcategoria").all()
    if fornecedor_id:
        qs = qs.filter(cdFornecedor_id=fornecedor_id)
    if categoria_id:
        qs = qs.filter(cdCategoria_id=categoria_id)
    if subcategoria_id:
        qs = qs.filter(cdSubcategoria_id=subcategoria_id)
    if inicio:
        try:
            inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
            qs = qs.filter(dtVencimento__gte=inicio_dt)
        except ValueError:
            pass
    if fim:
        try:
            fim_dt = datetime.strptime(fim, "%Y-%m-%d").date()
            qs = qs.filter(dtVencimento__lte=fim_dt)
        except ValueError:
            pass
    if status:
        if status == "ATRASADO":
            qs = qs.filter(dtVencimento__lt=today).exclude(status__in=["PAGO", "CANCELADO"])
        else:
            qs = qs.filter(status=status)

    qs = qs.order_by("dtVencimento", "id")
    paginator = Paginator(qs, 10)
    page = paginator.get_page(request.GET.get("page"))
    edit_forms = {obj.id: forms.ContasPagarForm(instance=obj) for obj in page}
    for obj in page:
        if obj.status == "PAGO":
            display_status = "PAGO"
        elif obj.status == "CANCELADO":
            display_status = "CANCELADO"
        elif obj.dtVencimento and obj.dtVencimento < today:
            display_status = "ATRASADO"
        else:
            display_status = "AGENDADO"
        obj.display_status = display_status

    query_params = request.GET.copy()
    if query_params.get("page"):
        query_params.pop("page")
    context = {
        "title": "Contas a Pagar",
        "page": page,
        "form": forms.ContasPagarForm(),
        "edit_forms": edit_forms,
        "fornecedores": models.Fornecedor.objects.all(),
        "categorias": models.Categoria.objects.all(),
        "subcategorias": models.Subcategoria.objects.all(),
        "filtros": {
            "inicio": inicio,
            "fim": fim,
            "status": status,
            "fornecedor": fornecedor_id,
            "categoria": categoria_id,
            "subcategoria": subcategoria_id,
        },
        "today": today.strftime("%Y-%m-%d"),
        "pagination_query": query_params.urlencode(),
        "breadcrumbs": [("Home", reverse("dashboard")), ("Contas a Pagar", "#")],
        "active_menu": "financeiro",
    }
    return render(request, "financeiro/contas_pagar_list.html", context)


def _filtrar_contas_pagar(request):
    today = timezone.now().date()
    inicio = request.GET.get("inicio", "").strip()
    fim = request.GET.get("fim", "").strip()
    status = request.GET.get("status", "").strip()
    fornecedor_id = request.GET.get("fornecedor", "").strip()
    categoria_id = request.GET.get("categoria", "").strip()
    subcategoria_id = request.GET.get("subcategoria", "").strip()

    qs = models.ContasPagar.objects.select_related("cdFornecedor", "cdCategoria", "cdSubcategoria").all()
    if fornecedor_id:
        qs = qs.filter(cdFornecedor_id=fornecedor_id)
    if categoria_id:
        qs = qs.filter(cdCategoria_id=categoria_id)
    if subcategoria_id:
        qs = qs.filter(cdSubcategoria_id=subcategoria_id)
    if inicio:
        try:
            inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
            qs = qs.filter(dtVencimento__gte=inicio_dt)
        except ValueError:
            pass
    if fim:
        try:
            fim_dt = datetime.strptime(fim, "%Y-%m-%d").date()
            qs = qs.filter(dtVencimento__lte=fim_dt)
        except ValueError:
            pass
    if status:
        if status == "ATRASADO":
            qs = qs.filter(dtVencimento__lt=today).exclude(status__in=["PAGO", "CANCELADO"])
        else:
            qs = qs.filter(status=status)
    return qs.order_by("dtVencimento", "id")


@login_required
def conta_bancaria_view(request):
    today = timezone.now().date()
    inicio = request.GET.get("inicio", "").strip()
    fim = request.GET.get("fim", "").strip()
    conta_id = request.GET.get("conta", "").strip()
    tipo = request.GET.get("tipo", "").strip()
    if not inicio or not fim:
        first, last = _first_last_day_month(today)
        if not inicio:
            inicio = first.strftime("%Y-%m-%d")
        if not fim:
            fim = last.strftime("%Y-%m-%d")

    contas = models.ContaBancaria.objects.filter(ativo=True).order_by("banco")
    conta_selecionada = contas.filter(pk=conta_id).first() if conta_id else contas.first()
    movimentos_qs = models.MovimentoConta.objects.select_related("conta").all()
    if conta_selecionada:
        movimentos_qs = movimentos_qs.filter(conta=conta_selecionada)
    if tipo:
        movimentos_qs = movimentos_qs.filter(tipo=tipo)
    if inicio:
        try:
            inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
            movimentos_qs = movimentos_qs.filter(data__gte=inicio_dt)
        except ValueError:
            inicio_dt = None
    else:
        inicio_dt = None
    if fim:
        try:
            fim_dt = datetime.strptime(fim, "%Y-%m-%d").date()
            movimentos_qs = movimentos_qs.filter(data__lte=fim_dt)
        except ValueError:
            fim_dt = None
    else:
        fim_dt = None
    movimentos = movimentos_qs.order_by("-data", "-id")
    total_entrada = movimentos_qs.filter(tipo="ENTRADA").aggregate(total=Sum("valor"))["total"] or 0
    total_saida = movimentos_qs.filter(tipo="SAIDA").aggregate(total=Sum("valor"))["total"] or 0
    saldo_inicial = conta_selecionada.saldo_inicial if conta_selecionada else 0
    saldo_atual = saldo_inicial + total_entrada - total_saida

    context = {
        "title": "Conta Bancaria",
        "contas": contas,
        "conta_selecionada": conta_selecionada,
        "movimentos": movimentos,
        "form_conta": forms.ContaBancariaForm(),
        "form_movimento": forms.MovimentoContaForm(),
        "filtros": {"inicio": inicio, "fim": fim, "conta": conta_id, "tipo": tipo},
        "total_entrada": total_entrada,
        "total_saida": total_saida,
        "saldo_inicial": saldo_inicial,
        "saldo_atual": saldo_atual,
        "today": today.strftime("%Y-%m-%d"),
        "breadcrumbs": [("Home", reverse("dashboard")), ("Conta Bancaria", "#")],
        "active_menu": "financeiro",
    }
    return render(request, "financeiro/conta_bancaria.html", context)


@login_required
def criar_conta_bancaria(request):
    if request.method != "POST":
        return redirect("conta_bancaria")
    data = request.POST.copy()
    data = _inject_cd_value(models.ContaBancaria, data)
    form = forms.ContaBancariaForm(data)
    if form.is_valid():
        form.save()
        messages.success(request, "Conta bancaria criada.")
    else:
        messages.error(request, "Verifique os erros.")
    return redirect("conta_bancaria")


@login_required
def criar_movimento_conta(request):
    if request.method != "POST":
        return redirect("conta_bancaria")
    data = request.POST.copy()
    form = forms.MovimentoContaForm(data, files=request.FILES or None)
    if form.is_valid():
        form.save()
        messages.success(request, "Lancamento registrado.")
    else:
        messages.error(request, "Verifique os erros.")
    return redirect("conta_bancaria")


@login_required
def dre_view(request):
    today = timezone.now().date()
    inicio = request.GET.get("inicio", "").strip()
    fim = request.GET.get("fim", "").strip()
    if not inicio or not fim:
        first, last = _first_last_day_month(today)
        if not inicio:
            inicio = first.strftime("%Y-%m-%d")
        if not fim:
            fim = last.strftime("%Y-%m-%d")
    try:
        inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
    except ValueError:
        inicio_dt = today.replace(day=1)
    try:
        fim_dt = datetime.strptime(fim, "%Y-%m-%d").date()
    except ValueError:
        fim_dt = today

    receitas = models.ContasReceber.objects.filter(status="PAGO", dtPagamento__range=(inicio_dt, fim_dt))
    despesas = models.ContasPagar.objects.filter(status="PAGO", dtPagamento__range=(inicio_dt, fim_dt))
    total_receitas = receitas.aggregate(total=Sum("valor"))["total"] or 0
    total_despesas = despesas.aggregate(total=Sum("valor"))["total"] or 0
    resultado = total_receitas - total_despesas
    status_label = "Lucro" if resultado >= 0 else "Prejuizo"

    context = {
        "title": "DRE",
        "inicio": inicio,
        "fim": fim,
        "total_receitas": total_receitas,
        "total_despesas": total_despesas,
        "resultado": resultado,
        "status_label": status_label,
        "receitas": receitas.order_by("-dtPagamento")[:10],
        "despesas": despesas.order_by("-dtPagamento")[:10],
        "breadcrumbs": [("Home", reverse("dashboard")), ("DRE", "#")],
        "active_menu": "financeiro",
    }
    return render(request, "financeiro/dre.html", context)


@login_required
def dre_relatorio(request):
    today = timezone.now().date()
    inicio = request.GET.get("inicio", "").strip()
    fim = request.GET.get("fim", "").strip()
    if not inicio or not fim:
        first, last = _first_last_day_month(today)
        if not inicio:
            inicio = first.strftime("%Y-%m-%d")
        if not fim:
            fim = last.strftime("%Y-%m-%d")
    try:
        inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
    except ValueError:
        inicio_dt = today.replace(day=1)
    try:
        fim_dt = datetime.strptime(fim, "%Y-%m-%d").date()
    except ValueError:
        fim_dt = today

    prev_month_last = inicio_dt.replace(day=1) - timedelta(days=1)
    movimentos_prev = models.MovimentoConta.objects.filter(data__lte=prev_month_last)
    saldo_inicial_total = models.ContaBancaria.objects.filter(ativo=True).aggregate(total=Sum("saldo_inicial"))["total"] or 0
    entradas_prev = movimentos_prev.filter(tipo="ENTRADA").aggregate(total=Sum("valor"))["total"] or 0
    saidas_prev = movimentos_prev.filter(tipo="SAIDA").aggregate(total=Sum("valor"))["total"] or 0
    saldo_bancario_anterior = saldo_inicial_total + entradas_prev - saidas_prev

    receitas_qs = models.ContasReceber.objects.filter(status="PAGO", dtPagamento__range=(inicio_dt, fim_dt))
    despesas_qs = models.ContasPagar.objects.filter(status="PAGO", dtPagamento__range=(inicio_dt, fim_dt))

    receita_bruta = receitas_qs.aggregate(total=Sum("valor"))["total"] or 0
    deducoes = 0
    receita_liquida = receita_bruta - deducoes
    custo_direto = 0
    lucro_bruto = receita_liquida - custo_direto
    despesas_operacionais = despesas_qs.aggregate(total=Sum("valor"))["total"] or 0
    resultado_final = lucro_bruto - despesas_operacionais
    saldo_bancario_final = saldo_bancario_anterior + receita_bruta - despesas_operacionais

    receitas_por_plano = (
        receitas_qs.values("contrato__cdPlano__dsPlano")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    receitas_por_categoria = (
        receitas_qs.values("contrato__cdPlano__categoria_receita__dsCategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_categoria = (
        despesas_qs.values("cdCategoria__dsCategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_subcategoria = (
        despesas_qs.values("cdSubcategoria__dsSubcategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_subcategoria = (
        despesas_qs.values("cdSubcategoria__dsSubcategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_subcategoria = (
        despesas_qs.values("cdSubcategoria__dsSubcategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_subcategoria = (
        despesas_qs.values("cdSubcategoria__dsSubcategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_subcategoria = (
        despesas_qs.values("cdSubcategoria__dsSubcategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_subcategoria = (
        despesas_qs.values("cdSubcategoria__dsSubcategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    receitas_labels = [item["contrato__cdPlano__dsPlano"] or "Sem plano" for item in receitas_por_plano]
    receitas_values = [float(item["total"] or 0) for item in receitas_por_plano]
    receitas_cat_labels = [item["contrato__cdPlano__categoria_receita__dsCategoria"] or "Sem categoria" for item in receitas_por_categoria]
    receitas_cat_values = [float(item["total"] or 0) for item in receitas_por_categoria]
    despesas_labels = [item["cdCategoria__dsCategoria"] or "Sem categoria" for item in despesas_por_categoria]
    despesas_values = [float(item["total"] or 0) for item in despesas_por_categoria]
    despesas_sub_labels = [item["cdSubcategoria__dsSubcategoria"] or "Sem subcategoria" for item in despesas_por_subcategoria]
    despesas_sub_values = [float(item["total"] or 0) for item in despesas_por_subcategoria]

    context = {
        "inicio": inicio,
        "fim": fim,
        "receita_bruta": receita_bruta,
        "deducoes": deducoes,
        "receita_liquida": receita_liquida,
        "custo_direto": custo_direto,
        "lucro_bruto": lucro_bruto,
        "despesas_operacionais": despesas_operacionais,
        "resultado_final": resultado_final,
        "resultado_label": "Lucro" if resultado_final >= 0 else "Prejuizo",
        "saldo_bancario_anterior": saldo_bancario_anterior,
        "saldo_bancario_final": saldo_bancario_final,
        "receitas_por_plano": receitas_por_plano,
        "receitas_por_categoria": receitas_por_categoria,
        "despesas_por_categoria": despesas_por_categoria,
        "despesas_por_subcategoria": despesas_por_subcategoria,
        "chart_receitas_labels": json.dumps(receitas_labels),
        "chart_receitas_values": json.dumps(receitas_values),
        "chart_receitas_cat_labels": json.dumps(receitas_cat_labels),
        "chart_receitas_cat_values": json.dumps(receitas_cat_values),
        "chart_despesas_labels": json.dumps(despesas_labels),
        "chart_despesas_values": json.dumps(despesas_values),
        "chart_despesas_sub_labels": json.dumps(despesas_sub_labels),
        "chart_despesas_sub_values": json.dumps(despesas_sub_values),
        "breadcrumbs": [("Home", reverse("dashboard")), ("DRE", reverse("dre_view")), ("Relatorio", "#")],
        "active_menu": "financeiro",
    }
    return render(request, "financeiro/dre_relatorio.html", context)


@login_required
def exportar_dre_excel(request):
    today = timezone.now().date()
    inicio = request.GET.get("inicio", "").strip()
    fim = request.GET.get("fim", "").strip()
    if not inicio or not fim:
        first, last = _first_last_day_month(today)
        if not inicio:
            inicio = first.strftime("%Y-%m-%d")
        if not fim:
            fim = last.strftime("%Y-%m-%d")
    try:
        inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
    except ValueError:
        inicio_dt = today.replace(day=1)
    try:
        fim_dt = datetime.strptime(fim, "%Y-%m-%d").date()
    except ValueError:
        fim_dt = today

    receitas_qs = models.ContasReceber.objects.filter(status="PAGO", dtPagamento__range=(inicio_dt, fim_dt))
    despesas_qs = models.ContasPagar.objects.filter(status="PAGO", dtPagamento__range=(inicio_dt, fim_dt))
    receita_bruta = receitas_qs.aggregate(total=Sum("valor"))["total"] or 0
    deducoes = 0
    receita_liquida = receita_bruta - deducoes
    custo_direto = 0
    lucro_bruto = receita_liquida - custo_direto
    despesas_operacionais = despesas_qs.aggregate(total=Sum("valor"))["total"] or 0
    resultado_final = lucro_bruto - despesas_operacionais

    receitas_por_plano = (
        receitas_qs.values("contrato__cdPlano__dsPlano")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    receitas_por_categoria = (
        receitas_qs.values("contrato__cdPlano__categoria_receita__dsCategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_categoria = (
        despesas_qs.values("cdCategoria__dsCategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Periodo inicio", inicio])
    ws.append(["Periodo fim", fim])
    ws.append(["Receita bruta", float(receita_bruta)])
    ws.append(["Deducoes", float(deducoes)])
    ws.append(["Receita liquida", float(receita_liquida)])
    ws.append(["Custo direto", float(custo_direto)])
    ws.append(["Lucro bruto", float(lucro_bruto)])
    ws.append(["Despesas operacionais", float(despesas_operacionais)])
    ws.append(["Resultado final", float(resultado_final)])

    ws_receitas = wb.create_sheet("Receitas por plano")
    ws_receitas.append(["Plano", "Total"])
    for item in receitas_por_plano:
        ws_receitas.append([item["contrato__cdPlano__dsPlano"] or "Sem plano", float(item["total"] or 0)])

    ws_receitas_cat = wb.create_sheet("Receitas por categoria")
    ws_receitas_cat.append(["Categoria", "Total"])
    for item in receitas_por_categoria:
        ws_receitas_cat.append([item["contrato__cdPlano__categoria_receita__dsCategoria"] or "Sem categoria", float(item["total"] or 0)])

    ws_despesas = wb.create_sheet("Despesas por categoria")
    ws_despesas.append(["Categoria", "Total"])
    for item in despesas_por_categoria:
        ws_despesas.append([item["cdCategoria__dsCategoria"] or "Sem categoria", float(item["total"] or 0)])

    ws_despesas_sub = wb.create_sheet("Despesas por subcategoria")
    ws_despesas_sub.append(["Subcategoria", "Total"])
    for item in despesas_por_subcategoria:
        ws_despesas_sub.append([item["cdSubcategoria__dsSubcategoria"] or "Sem subcategoria", float(item["total"] or 0)])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="dre-completo.xlsx"'
    return response


@login_required
def exportar_dre_pdf(request):
    today = timezone.now().date()
    inicio = request.GET.get("inicio", "").strip()
    fim = request.GET.get("fim", "").strip()
    if not inicio or not fim:
        first, last = _first_last_day_month(today)
        if not inicio:
            inicio = first.strftime("%Y-%m-%d")
        if not fim:
            fim = last.strftime("%Y-%m-%d")
    try:
        inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
    except ValueError:
        inicio_dt = today.replace(day=1)
    try:
        fim_dt = datetime.strptime(fim, "%Y-%m-%d").date()
    except ValueError:
        fim_dt = today

    receitas_qs = models.ContasReceber.objects.filter(status="PAGO", dtPagamento__range=(inicio_dt, fim_dt))
    despesas_qs = models.ContasPagar.objects.filter(status="PAGO", dtPagamento__range=(inicio_dt, fim_dt))
    receita_bruta = receitas_qs.aggregate(total=Sum("valor"))["total"] or 0
    deducoes = 0
    receita_liquida = receita_bruta - deducoes
    custo_direto = 0
    lucro_bruto = receita_liquida - custo_direto
    despesas_operacionais = despesas_qs.aggregate(total=Sum("valor"))["total"] or 0
    resultado_final = lucro_bruto - despesas_operacionais

    receitas_por_plano = (
        receitas_qs.values("contrato__cdPlano__dsPlano")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    receitas_por_categoria = (
        receitas_qs.values("contrato__cdPlano__categoria_receita__dsCategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_categoria = (
        despesas_qs.values("cdCategoria__dsCategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_subcategoria = (
        despesas_qs.values("cdSubcategoria__dsSubcategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="DRE Completo")
    styles = getSampleStyleSheet()
    title = Paragraph("DRE Completo", styles["Title"])
    subtitle = Paragraph(f"Periodo {inicio_dt.strftime('%d/%m/%Y')} a {fim_dt.strftime('%d/%m/%Y')}", styles["Normal"])
    resumo = [
        ["Receita bruta", f"R$ {receita_bruta}"],
        ["Deducoes", f"R$ {deducoes}"],
        ["Receita liquida", f"R$ {receita_liquida}"],
        ["Custo direto", f"R$ {custo_direto}"],
        ["Lucro bruto", f"R$ {lucro_bruto}"],
        ["Despesas operacionais", f"R$ {despesas_operacionais}"],
        ["Resultado final", f"R$ {resultado_final}"],
    ]
    resumo_table = Table(resumo, colWidths=[220, 160])
    resumo_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d4d4d")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dbe4ea")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )

    receitas_table = [["Plano", "Total"]]
    for item in receitas_por_plano:
        receitas_table.append([item["contrato__cdPlano__dsPlano"] or "Sem plano", f"R$ {item['total']}"])
    receitas_table = Table(receitas_table, colWidths=[260, 120])
    receitas_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e0f2fe")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dbe4ea")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )

    receitas_cat_table = [["Categoria", "Total"]]
    for item in receitas_por_categoria:
        receitas_cat_table.append([item["contrato__cdPlano__categoria_receita__dsCategoria"] or "Sem categoria", f"R$ {item['total']}"])
    receitas_cat_table = Table(receitas_cat_table, colWidths=[260, 120])
    receitas_cat_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dcfce7")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dbe4ea")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )

    despesas_table = [["Categoria", "Total"]]
    for item in despesas_por_categoria:
        despesas_table.append([item["cdCategoria__dsCategoria"] or "Sem categoria", f"R$ {item['total']}"])
    despesas_table = Table(despesas_table, colWidths=[260, 120])
    despesas_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fee2e2")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dbe4ea")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )

    elements = [
        title,
        subtitle,
        Spacer(1, 12),
        resumo_table,
        Spacer(1, 16),
        Paragraph("Receitas por plano", styles["Heading3"]),
        receitas_table,
        Spacer(1, 16),
        Paragraph("Receitas por categoria", styles["Heading3"]),
        receitas_cat_table,
        Spacer(1, 16),
        Paragraph("Despesas por categoria", styles["Heading3"]),
        despesas_table,
        Spacer(1, 16),
        Paragraph("Despesas por subcategoria", styles["Heading3"]),
        Table(
            [["Subcategoria", "Total"]] + [
                [item["cdSubcategoria__dsSubcategoria"] or "Sem subcategoria", f"R$ {item['total']}"]
                for item in despesas_por_subcategoria
            ],
            colWidths=[260, 120],
        ),
    ]
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="dre-completo.pdf"'
    return response

@login_required
def exportar_contas_pagar_excel(request):
    qs = _filtrar_contas_pagar(request)
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Contas a Pagar"
    ws.append(["Fornecedor", "Categoria", "Subcategoria", "Vencimento", "Pagamento", "Valor", "Status"])
    for f in qs:
        if f.status == "PAGO":
            display_status = "PAGO"
        elif f.status == "CANCELADO":
            display_status = "CANCELADO"
        elif f.dtVencimento and f.dtVencimento < timezone.now().date():
            display_status = "ATRASADO"
        else:
            display_status = "AGENDADO"
        ws.append(
            [
                str(f.cdFornecedor),
                str(f.cdCategoria),
                str(f.cdSubcategoria),
                f.dtVencimento.strftime("%d/%m/%Y") if f.dtVencimento else "",
                f.dtPagamento.strftime("%d/%m/%Y") if f.dtPagamento else "",
                float(f.valor),
                display_status,
            ]
        )
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="contas-a-pagar.xlsx"'
    return response


@login_required
def exportar_contas_pagar_pdf(request):
    qs = _filtrar_contas_pagar(request)
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="Contas a Pagar")
    styles = getSampleStyleSheet()
    title = Paragraph("Contas a Pagar", styles["Title"])
    data = [["Fornecedor", "Categoria", "Subcategoria", "Vencimento", "Pagamento", "Valor", "Status"]]
    for f in qs:
        if f.status == "PAGO":
            display_status = "PAGO"
        elif f.status == "CANCELADO":
            display_status = "CANCELADO"
        elif f.dtVencimento and f.dtVencimento < timezone.now().date():
            display_status = "ATRASADO"
        else:
            display_status = "AGENDADO"
        data.append(
            [
                str(f.cdFornecedor),
                str(f.cdCategoria),
                str(f.cdSubcategoria),
                f.dtVencimento.strftime("%d/%m/%Y") if f.dtVencimento else "-",
                f.dtPagamento.strftime("%d/%m/%Y") if f.dtPagamento else "-",
                f"R$ {f.valor}",
                display_status,
            ]
        )
    table = Table(data, repeatRows=1, colWidths=[90, 80, 90, 70, 70, 60, 70])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d4d4d")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dbe4ea")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f7fb")]),
            ]
        )
    )
    elements = [title, Spacer(1, 12), table]
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="contas-a-pagar.pdf"'
    return response


@login_required
def aulas_list(request):
    qs = models.AulaSessao.objects.select_related("unidade", "tipoServico", "profissional")
    week_str = request.GET.get("week", "").strip()
    profissional_id = request.GET.get("profissional", "").strip()
    view_mode = request.GET.get("view", "week").strip().lower()
    if view_mode == "list":
        view_mode = "week"
    if view_mode not in {"week", "day", "month"}:
        view_mode = "week"
    try:
        ref_date = datetime.strptime(week_str, "%Y-%m-%d").date() if week_str else date.today()
    except ValueError:
        ref_date = date.today()

    if view_mode == "day":
        week_start = ref_date
        week_end = ref_date
        days = [ref_date]
        columns = 1
    elif view_mode == "month":
        first = ref_date.replace(day=1)
        if first.month == 12:
            last = date(first.year + 1, 1, 1) - timedelta(days=1)
        else:
            last = date(first.year, first.month + 1, 1) - timedelta(days=1)
        week_start = first
        week_end = last
        leading = first.weekday()
        days = [None] * leading + [first + timedelta(days=i) for i in range((last - first).days + 1)]
        columns = 7
    else:
        week_start = ref_date - timedelta(days=ref_date.weekday())
        week_end = week_start + timedelta(days=5)
        days = [week_start + timedelta(days=i) for i in range(6)]
        columns = 6

    qs = qs.filter(data__range=(week_start, week_end))
    if profissional_id:
        qs = qs.filter(profissional_id=profissional_id)

    aulas = list(qs.order_by("data", "horaInicio"))
    aula_ids = [aula.id for aula in aulas]
    reservas = (
        models.Reserva.objects.filter(aulaSessao_id__in=aula_ids)
        .select_related("aluno", "aulaSessao", "aulaSessao__profissional")
        .order_by("aulaSessao__data", "aulaSessao__horaInicio")
    )
    reservas_by_aula = {}
    for reserva in reservas:
        reservas_by_aula.setdefault(reserva.aulaSessao_id, []).append(reserva)

    aulas_by_day = {day: [] for day in days if day}
    for aula in aulas:
        aulas_by_day.setdefault(aula.data, []).append(aula)
    for aulas in aulas_by_day.values():
        aulas.sort(key=lambda x: x.horaInicio)

    weekday_labels = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sab", "Dom"]
    week_cards = []
    for dia in days:
        if not dia:
            week_cards.append({"is_placeholder": True})
            continue
        week_cards.append(
            {
                "date": dia,
                "label": weekday_labels[dia.weekday()],
                "number": dia.day,
                "is_today": dia == date.today(),
                "aulas": aulas_by_day.get(dia, []),
            }
        )

    prof_chips = []
    for prof in models.Profissional.objects.all():
        parts = [p for p in prof.profissional.split() if p]
        initials = "".join([p[0] for p in parts[:2]]).upper()
        prof_chips.append({"name": prof.profissional, "initials": initials})

    meses = [
        "Janeiro",
        "Fevereiro",
        "Marco",
        "Abril",
        "Maio",
        "Junho",
        "Julho",
        "Agosto",
        "Setembro",
        "Outubro",
        "Novembro",
        "Dezembro",
    ]
    month_label = f"{meses[week_start.month - 1]} {week_start.year}"

    context = {
        "title": "Aulas",
        "week_cards": week_cards,
        "week_start": week_start,
        "week_end": week_end,
        "prev_week": week_start - timedelta(days=7),
        "next_week": week_start + timedelta(days=7),
        "month_label": month_label,
        "view_mode": view_mode,
        "columns": columns,
        "form": forms.AulaSessaoForm(),
        "profissionais": models.Profissional.objects.all(),
        "prof_chips": prof_chips[:5],
        "reservas_by_aula": reservas_by_aula,
        "modelos_evolucao": models.ModeloEvolucao.objects.filter(ativo=True).order_by("titulo"),
        "breadcrumbs": [("Home", reverse("dashboard")), ("Aulas", "#")],
        "active_menu": "agenda",
    }
    return render(request, "agenda/aulas_list.html", context)


def create_view(request, model, form_class, redirect_name):
    if request.method == "POST":
        data = request.POST.copy()
        data = _inject_cd_value(model, data)
        if model is models.AulaSessao and not data.get("horaFim"):
            unidade_id = data.get("unidade")
            hora_inicio = data.get("horaInicio")
            if unidade_id and hora_inicio:
                unidade = models.Unidade.objects.filter(pk=unidade_id).first()
                if unidade and unidade.duracao_aula_minutos:
                    inicio_time = datetime.strptime(hora_inicio, "%H:%M").time()
                    base = datetime.combine(date.today(), inicio_time)
                    fim = base + timedelta(minutes=unidade.duracao_aula_minutos)
                    data["horaFim"] = fim.time().strftime("%H:%M")
        if model is models.Profissional and not data.get("cdPerfilAcesso"):
            perfil, _ = models.PerfilAcesso.objects.get_or_create(
                cdPerfilAcesso=1, defaults={"dsPerfilAcesso": "Padrao"}
            )
            data["cdPerfilAcesso"] = perfil.id
        form = form_class(data, files=request.FILES or None)
        if form.is_valid():
            if model is models.Contrato:
                cleaned = form.cleaned_data
                valor_parcela = cleaned.get("valor_parcela") or float(data.get("valor", "0") or 0)
                valor_total = cleaned.get("valor_total") or 0
                if not valor_total and cleaned.get("cdPlano"):
                    valor_total = float(valor_parcela) * float(cleaned["cdPlano"].duracao_meses or 1)
                contrato_data = {
                    "cdContrato": cleaned["cdContrato"],
                    "cdAluno": cleaned["cdAluno"],
                    "cdPlano": cleaned["cdPlano"],
                    "cdUnidade": cleaned["cdUnidade"],
                    "cdProfissional": cleaned["cdProfissional"],
                    "valor_parcela": valor_parcela,
                    "valor_total": valor_total,
                    "dtInicioContrato": cleaned["dtInicioContrato"],
                    "dtFimContrato": cleaned["dtFimContrato"],
                }
                valor = float(valor_parcela or 0)
                obj = services.criar_contrato_e_contas(contrato_data, valor)
                if services.enviar_contrato_para_assinatura(obj, request.build_absolute_uri("/")):
                    messages.success(request, "Contrato criado e enviado por email. Agende as aulas.")
                else:
                    messages.warning(request, "Contrato criado, mas aluno sem email para assinatura.")
                    messages.success(request, "Contrato criado. Agende as aulas.")
                return redirect("contratos_agenda", pk=obj.id)
            obj = form.save()
            if model is models.ContasPagar:
                recorrencia = obj.recorrencia
                quantidade = obj.recorrencia_quantidade or 0
                if recorrencia and quantidade < 1:
                    quantidade = 1
                    obj.recorrencia_quantidade = quantidade
                    obj.save(update_fields=["recorrencia_quantidade"])
                if recorrencia and quantidade > 1:
                    max_cd = models.ContasPagar.objects.order_by("-cdContasPagar").values_list("cdContasPagar", flat=True).first() or 0
                    for idx in range(1, quantidade):
                        if recorrencia == "SEMANAL":
                            vencimento = obj.dtVencimento + timedelta(days=7 * idx)
                        elif recorrencia == "ANUAL":
                            vencimento = _add_years(obj.dtVencimento, idx)
                        else:
                            vencimento = _add_months(obj.dtVencimento, idx)
                        max_cd += 1
                        models.ContasPagar.objects.create(
                            cdContasPagar=max_cd,
                            cdFornecedor=obj.cdFornecedor,
                            cdCategoria=obj.cdCategoria,
                            cdSubcategoria=obj.cdSubcategoria,
                            dtVencimento=vencimento,
                            valor=obj.valor,
                            recorrencia=obj.recorrencia,
                            recorrencia_quantidade=obj.recorrencia_quantidade,
                        )
            if model is models.Profissional:
                _sync_user_for_profissional(obj, raw_password=form.cleaned_data.get("password"))
            if model is models.Aluno:
                _sync_aluno_address(obj, request.POST)
                _sync_aluno_phones(obj, request.POST)
            messages.success(request, "Salvo com sucesso")
        else:
            messages.error(request, "Verifique os erros")
        return redirect(redirect_name)
    return render(
        request,
        "generic/form.html",
        {"form": form_class(), "title": "Novo", "model_name": model._meta.model_name, "active_menu": _active_menu(request.path)},
    )


def edit_view(request, model, form_class, redirect_name, pk):
    obj = get_object_or_404(model, pk=pk)
    if request.method == "POST":
        next_url = request.POST.get("next", "").strip()
        if next_url and not next_url.startswith("/"):
            next_url = ""
        data = request.POST.copy()
        for field in model._meta.fields:
            if field.name.startswith("cd") and not data.get(field.name):
                data[field.name] = getattr(obj, field.name)
        if model is models.AulaSessao and not data.get("horaFim"):
            unidade_id = data.get("unidade")
            hora_inicio = data.get("horaInicio")
            if unidade_id and hora_inicio:
                unidade = models.Unidade.objects.filter(pk=unidade_id).first()
                if unidade and unidade.duracao_aula_minutos:
                    inicio_time = datetime.strptime(hora_inicio, "%H:%M").time()
                    base = datetime.combine(date.today(), inicio_time)
                    fim = base + timedelta(minutes=unidade.duracao_aula_minutos)
                    data["horaFim"] = fim.time().strftime("%H:%M")
        form = form_class(data, files=request.FILES or None, instance=obj)
        if form.is_valid():
            obj = form.save()
            if model is models.Profissional:
                _sync_user_for_profissional(obj, raw_password=form.cleaned_data.get("password"))
            if model is models.Aluno:
                _sync_aluno_address(obj, request.POST)
                _sync_aluno_phones(obj, request.POST)
            if model is models.Contrato and obj.status == "NAO_ASSINADO":
                if not services.enviar_contrato_para_assinatura(obj, request.build_absolute_uri("/")):
                    messages.warning(request, "Contrato atualizado, mas aluno sem email para assinatura.")
            messages.success(request, "Atualizado com sucesso")
            return redirect(next_url or redirect_name)
        messages.error(request, "Verifique os erros")
        if next_url:
            return redirect(next_url)
    return render(
        request,
        "generic/form.html",
        {"form": form_class(instance=obj), "title": "Editar", "model_name": model._meta.model_name, "active_menu": _active_menu(request.path)},
    )


def delete_view(request, model, redirect_name, pk):
    obj = get_object_or_404(model, pk=pk)
    if request.method == "POST":
        next_url = request.POST.get("next", "").strip()
        if next_url and not next_url.startswith("/"):
            next_url = ""
        if model is models.Contrato:
            models.ContasReceber.objects.filter(contrato=obj, status="ABERTO").delete()
            if models.ContasReceber.objects.filter(contrato=obj).exists():
                messages.error(request, "Nao foi possivel excluir: existem contas pagas ou atrasadas.")
                return redirect(next_url or redirect_name)
        obj.delete()
        messages.success(request, "Removido")
        return redirect(next_url or redirect_name)
    return render(
        request,
        "generic/confirm_delete.html",
        {"object": obj, "title": "Excluir", "active_menu": _active_menu(request.path)},
    )


@login_required
def baixar_conta_receber(request, pk):
    conta = get_object_or_404(models.ContasReceber, pk=pk)
    if request.method == "POST":
        next_url = request.POST.get("next", "").strip()
        if next_url and not next_url.startswith("/"):
            next_url = ""
        data_pagamento = request.POST.get("dtPagamento") or ""
        try:
            pago_em = datetime.strptime(data_pagamento, "%Y-%m-%d").date() if data_pagamento else timezone.now().date()
        except ValueError:
            pago_em = timezone.now().date()
        conta.status = "PAGO"
        conta.dtPagamento = pago_em
        conta.save(update_fields=["status", "dtPagamento"])
        messages.success(request, "Lancamento baixado com sucesso.")
        return redirect(next_url or "contas_receber_list")
    return redirect("contas_receber_list")


@login_required
def efetuar_pagamento_conta_pagar(request, pk):
    conta = get_object_or_404(models.ContasPagar, pk=pk)
    if request.method != "POST":
        return redirect("contas_pagar_list")
    next_url = request.POST.get("next", "").strip()
    if next_url and not next_url.startswith("/"):
        next_url = ""
    data_pagamento = request.POST.get("dtPagamento") or ""
    try:
        pago_em = datetime.strptime(data_pagamento, "%Y-%m-%d").date() if data_pagamento else timezone.now().date()
    except ValueError:
        pago_em = timezone.now().date()
    comprovante = request.FILES.get("comprovante")
    conta.status = "PAGO"
    conta.dtPagamento = pago_em
    if comprovante:
        conta.comprovante = comprovante
    conta.save(update_fields=["status", "dtPagamento", "comprovante"])
    messages.success(request, "Pagamento registrado.")
    return redirect(next_url or "contas_pagar_list")


@login_required
def cancelar_conta_pagar(request, pk):
    conta = get_object_or_404(models.ContasPagar, pk=pk)
    if request.method != "POST":
        return redirect("contas_pagar_list")
    next_url = request.POST.get("next", "").strip()
    if next_url and not next_url.startswith("/"):
        next_url = ""
    motivo = request.POST.get("motivo_cancelamento", "").strip()
    conta.status = "CANCELADO"
    conta.motivo_cancelamento = motivo
    conta.save(update_fields=["status", "motivo_cancelamento"])
    messages.success(request, "Lancamento cancelado.")
    return redirect(next_url or "contas_pagar_list")


@login_required
def evoluir_reserva(request, pk):
    reserva = get_object_or_404(models.Reserva, pk=pk)
    if request.method != "POST":
        return redirect("aulas_list")
    next_url = request.POST.get("next", "").strip()
    if next_url and not next_url.startswith("/"):
        next_url = ""
    status = request.POST.get("status", "").strip()
    texto = request.POST.get("texto", "").strip()
    if status not in {"CONCLUIDA", "FALTOU_AVISOU", "FALTOU_SEM_AVISAR"}:
        messages.error(request, "Status invalido.")
        return redirect(next_url or "aulas_list")
    if not texto:
        messages.error(request, "Preencha a evolucao.")
        return redirect(next_url or "aulas_list")
    if not reserva.aulaSessao.profissional_id:
        messages.error(request, "Defina o profissional da aula antes de registrar a evolucao.")
        return redirect(next_url or "aulas_list")
    models.EvolucaoAluno.objects.create(
        reserva=reserva,
        profissional=reserva.aulaSessao.profissional,
        texto=texto,
    )
    reserva.status = status
    reserva.save(update_fields=["status"])
    messages.success(request, "Evolucao registrada.")
    return redirect(next_url or "aulas_list")


def _get_filtros_financeiro(request):
    return {
        "status": request.GET.get("status", "").strip(),
        "inicio": request.GET.get("inicio", "").strip(),
        "fim": request.GET.get("fim", "").strip(),
    }


def _filtrar_contas_receber(qs, request):
    filtros = _get_filtros_financeiro(request)
    if filtros["status"]:
        qs = qs.filter(status=filtros["status"])
    inicio = filtros["inicio"]
    fim = filtros["fim"]
    if inicio:
        try:
            inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
            qs = qs.filter(dtVencimento__gte=inicio_dt)
        except ValueError:
            pass
    if fim:
        try:
            fim_dt = datetime.strptime(fim, "%Y-%m-%d").date()
            qs = qs.filter(dtVencimento__lte=fim_dt)
        except ValueError:
            pass
    return qs.order_by("dtVencimento", "id")


@login_required
def exportar_contas_receber_excel(request, aluno_id):
    aluno = get_object_or_404(models.Aluno, pk=aluno_id)
    qs = _filtrar_contas_receber(
        models.ContasReceber.objects.filter(contrato__cdAluno=aluno).select_related("contrato"),
        request,
    )
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Faturas"
    ws.append(["Competencia", "Vencimento", "Pagamento", "Contrato", "Status", "Valor"])
    for f in qs:
        ws.append(
            [
                f.competencia or "",
                f.dtVencimento.strftime("%d/%m/%Y") if f.dtVencimento else "",
                f.dtPagamento.strftime("%d/%m/%Y") if f.dtPagamento else "",
                f.contrato.cdContrato if f.contrato_id else "",
                f.status,
                float(f.valor),
            ]
        )
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="faturas-aluno-{aluno.id}.xlsx"'
    return response


@login_required
def exportar_contas_receber_pdf(request, aluno_id):
    aluno = get_object_or_404(models.Aluno, pk=aluno_id)
    qs = _filtrar_contas_receber(
        models.ContasReceber.objects.filter(contrato__cdAluno=aluno).select_related("contrato"),
        request,
    )
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="Faturas do Aluno")
    styles = getSampleStyleSheet()
    title = Paragraph(f"Faturas do aluno: {aluno.dsNome}", styles["Title"])
    subtitle = Paragraph("Resumo financeiro", styles["Normal"])
    data = [["Competencia", "Vencimento", "Pagamento", "Contrato", "Status", "Valor"]]
    for f in qs:
        data.append(
            [
                f.competencia or "-",
                f.dtVencimento.strftime("%d/%m/%Y") if f.dtVencimento else "-",
                f.dtPagamento.strftime("%d/%m/%Y") if f.dtPagamento else "-",
                str(f.contrato.cdContrato) if f.contrato_id else "-",
                f.status,
                f"R$ {f.valor}",
            ]
        )
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d4d4d")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dbe4ea")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f7fb")]),
                ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
            ]
        )
    )
    elements = [title, subtitle, Spacer(1, 12), table]
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="faturas-aluno-{aluno.id}.pdf"'
    return response


@login_required
def recibo_conta_receber_pdf(request, pk):
    conta = get_object_or_404(models.ContasReceber, pk=pk)
    aluno = conta.contrato.cdAluno if conta.contrato_id else None
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="Recibo")
    styles = getSampleStyleSheet()
    title = Paragraph("Recibo de Pagamento", styles["Title"])
    aluno_nome = aluno.dsNome if aluno else "-"
    pago_em = conta.dtPagamento.strftime("%d/%m/%Y") if conta.dtPagamento else "-"
    data = [
        ["Aluno", aluno_nome],
        ["Contrato", str(conta.contrato.cdContrato) if conta.contrato_id else "-"],
        ["Competencia", conta.competencia or "-"],
        ["Vencimento", conta.dtVencimento.strftime("%d/%m/%Y") if conta.dtVencimento else "-"],
        ["Pagamento", pago_em],
        ["Valor", f"R$ {conta.valor}"],
        ["Status", conta.status],
    ]
    table = Table(data, colWidths=[120, 360])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d4d4d")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dbe4ea")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f7fb")]),
            ]
        )
    )
    elements = [title, Spacer(1, 8), table, Spacer(1, 18), Paragraph("Obrigado!", styles["Heading3"])]
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="recibo-{conta.id}.pdf"'
    return response


@login_required
def exportar_evolucoes_excel(request, aluno_id):
    aluno = get_object_or_404(models.Aluno, pk=aluno_id)
    qs = (
        models.EvolucaoAluno.objects.filter(reserva__aluno=aluno)
        .select_related("profissional", "reserva")
        .order_by("-dtEvolucao")
    )
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Evolucoes"
    ws.append(["Data", "Profissional", "Evolucao"])
    for e in qs:
        ws.append(
            [
                e.dtEvolucao.strftime("%d/%m/%Y %H:%M") if e.dtEvolucao else "",
                str(e.profissional),
                e.texto,
            ]
        )
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="evolucoes-aluno-{aluno.id}.xlsx"'
    return response


@login_required
def exportar_evolucoes_pdf(request, aluno_id):
    aluno = get_object_or_404(models.Aluno, pk=aluno_id)
    qs = (
        models.EvolucaoAluno.objects.filter(reserva__aluno=aluno)
        .select_related("profissional", "reserva")
        .order_by("-dtEvolucao")
    )
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="Evolucoes do Aluno")
    styles = getSampleStyleSheet()
    title = Paragraph(f"Evolucoes do aluno: {aluno.dsNome}", styles["Title"])
    data = [["Data", "Profissional", "Evolucao"]]
    for e in qs:
        data.append(
            [
                e.dtEvolucao.strftime("%d/%m/%Y %H:%M") if e.dtEvolucao else "-",
                str(e.profissional),
                e.texto,
            ]
        )
    table = Table(data, repeatRows=1, colWidths=[110, 140, 290])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d4d4d")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dbe4ea")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f7fb")]),
            ]
        )
    )
    elements = [title, Spacer(1, 12), table]
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="evolucoes-aluno-{aluno.id}.pdf"'
    return response


@login_required
def wizard_step1(request):
    if request.method == "POST":
        if request.FILES.get("documento"):
            data = extract_student_from_document(request.FILES["documento"].read(), request.FILES["documento"].name)
            request.session["wizard_student"] = data
            return render(
                request,
                "wizard/step1_documento.html",
                {
                    "student": data,
                    "unidades": models.Unidade.objects.all(),
                    "breadcrumbs": [("Home", reverse("dashboard")), ("Wizard", "#")],
                    "active_menu": "cadastros",
                },
            )
        if request.POST.get("confirm") == "1":
            cpf = request.POST.get("cpf", "")
            if models.Aluno.objects.filter(dsCPF=cpf).exists():
                messages.error(request, "CPF duplicado")
                return redirect("wizard_step1")
            max_cd = models.Aluno.objects.order_by("-cdAluno").values_list("cdAluno", flat=True).first() or 0
            aluno = models.Aluno.objects.create(
                cdAluno=max_cd + 1,
                dsNome=request.POST.get("nome", ""),
                dsCPF=cpf,
                dsRg=request.POST.get("rg", ""),
                cdUnidade_id=int(request.POST.get("cdUnidade")),
            )
            request.session["wizard_aluno_id"] = aluno.id
            return redirect("wizard_step2")
    return render(
        request,
        "wizard/step1_documento.html",
        {"unidades": models.Unidade.objects.all(), "breadcrumbs": [("Home", reverse("dashboard")), ("Wizard", "#")], "active_menu": "cadastros"},
    )


@login_required
def wizard_step2(request):
    if request.method == "POST":
        if request.FILES.get("comprovante"):
            data = extract_address_from_proof(request.FILES["comprovante"].read(), request.FILES["comprovante"].name)
            request.session["wizard_address"] = data
            return render(
                request,
                "wizard/step2_endereco.html",
                {"address": data, "breadcrumbs": [("Home", reverse("dashboard")), ("Wizard", "#")], "active_menu": "cadastros"},
            )
        if request.POST.get("confirm") == "1":
            aluno_id = request.session.get("wizard_aluno_id")
            if not aluno_id:
                messages.error(request, "Aluno nao encontrado")
                return redirect("wizard_step1")
            max_cd = models.EnderecoAluno.objects.order_by("-cdEndereco").values_list("cdEndereco", flat=True).first() or 0
            endereco = models.EnderecoAluno.objects.create(
                cdEndereco=max_cd + 1,
                cdAluno_id=aluno_id,
                dsLogradouro=request.POST.get("logradouro", ""),
                dsNumero=request.POST.get("numero", ""),
                dsCEP=request.POST.get("cep", ""),
                dsCidade=request.POST.get("cidade", ""),
                dsBairro=request.POST.get("bairro", ""),
            )
            models.Aluno.objects.filter(pk=aluno_id).update(cdEndereco=endereco)
            return redirect("wizard_step3")
    return render(
        request,
        "wizard/step2_endereco.html",
        {"breadcrumbs": [("Home", reverse("dashboard")), ("Wizard", "#")], "active_menu": "cadastros"},
    )


@login_required
def wizard_step3(request):
    termos = models.TermoUso.objects.all()
    if request.method == "POST":
        termo_id = request.POST.get("termo")
        aluno_id = request.session.get("wizard_aluno_id")
        if aluno_id and termo_id:
            services.registrar_aceite_termo(models.Aluno.objects.get(pk=aluno_id), models.TermoUso.objects.get(pk=int(termo_id)))
        return redirect("wizard_step4")
    return render(
        request,
        "wizard/step3_termo.html",
        {"termos": termos, "breadcrumbs": [("Home", reverse("dashboard")), ("Wizard", "#")], "active_menu": "cadastros"},
    )


@login_required
def wizard_step4(request):
    if request.method == "POST":
        request.session["wizard_contrato"] = request.POST.dict()
        return redirect("wizard_step5")
    context = {
        "alunos": models.Aluno.objects.all(),
        "aluno_id": request.session.get("wizard_aluno_id"),
        "planos": models.Plano.objects.all(),
        "unidades": models.Unidade.objects.all(),
        "profissionais": models.Profissional.objects.all(),
        "breadcrumbs": [("Home", reverse("dashboard")), ("Wizard", "#")],
        "active_menu": "cadastros",
    }
    return render(request, "wizard/step4_contrato.html", context)


@login_required
def wizard_step5(request):
    if request.method == "POST":
        data = request.session.get("wizard_contrato", {})
        aluno = models.Aluno.objects.get(pk=int(data.get("cdAluno")))
        plano = models.Plano.objects.get(pk=int(data.get("cdPlano")))
        valor_parcela = float(data.get("valor_parcela") or data.get("valor") or 0)
        valor_total = float(data.get("valor_total") or 0)
        if not valor_total and plano:
            valor_total = float(valor_parcela) * float(plano.duracao_meses or 1)
        contrato_data = {
            "cdContrato": int(data.get("cdContrato")),
            "cdAluno": aluno,
            "cdPlano": plano,
            "cdUnidade": models.Unidade.objects.get(pk=int(data.get("cdUnidade"))),
            "cdProfissional": models.Profissional.objects.get(pk=int(data.get("cdProfissional"))),
            "valor_parcela": valor_parcela,
            "valor_total": valor_total,
            "dtInicioContrato": data.get("dtInicioContrato"),
            "dtFimContrato": data.get("dtFimContrato"),
        }
        contrato = services.criar_contrato_e_contas(contrato_data, float(data.get("valor", "0")))
        if services.enviar_contrato_para_assinatura(contrato, request.build_absolute_uri("/")):
            messages.success(request, "Contrato criado e enviado por email. Agende as aulas.")
        else:
            messages.warning(request, "Contrato criado, mas aluno sem email para assinatura.")
            messages.success(request, "Contrato criado. Agende as aulas.")
        return redirect("contratos_agenda", pk=contrato.id)
    return render(
        request,
        "wizard/step5_reservas.html",
        {"breadcrumbs": [("Home", reverse("dashboard")), ("Wizard", "#")], "active_menu": "cadastros"},
    )


@login_required
def contrato_agenda(request, pk):
    contrato = get_object_or_404(models.Contrato, pk=pk)
    plano = contrato.cdPlano
    aulas_por_semana = plano.aulas_por_semana or 1
    aulas = models.AulaSessao.objects.filter(
        unidade=contrato.cdUnidade,
        tipoServico=plano.cdTipoServico,
        data__range=(contrato.dtInicioContrato, contrato.dtFimContrato),
    ).select_related("unidade").order_by("data", "horaInicio")

    profissionais = list(models.Profissional.objects.all())
    prof_ids = [prof.id for prof in profissionais]
    horarios = models.HorarioStudio.objects.filter(
        unidade=contrato.cdUnidade,
        tipoServico=plano.cdTipoServico,
    ).order_by("diaSemana", "horaInicio")

    slots = {}
    aulas_by_key = {(aula.data, aula.horaInicio, aula.horaFim, aula.profissional_id): aula for aula in aulas}
    capacidade_padrao = contrato.cdUnidade.capacidade or 0
    horarios_by_slot = {}
    for horario in horarios:
        key = (horario.diaSemana, horario.horaInicio, horario.horaFim)
        horarios_by_slot.setdefault(key, []).append(horario)

    def _capacidade_para(horarios_list, prof_id):
        match = next((h for h in horarios_list if h.profissional_id == prof_id), None)
        if match and match.capacidade is not None:
            return match.capacidade
        general = next((h for h in horarios_list if h.profissional_id is None and h.capacidade is not None), None)
        if general:
            return general.capacidade
        return capacidade_padrao

    for key, horarios_list in horarios_by_slot.items():
        weekday, inicio, fim = key
        allowed = {h.profissional_id for h in horarios_list if h.profissional_id}
        candidate_profs = allowed if allowed else set(prof_ids)
        slot_ok = False
        for prof_id in candidate_profs:
            if not prof_id:
                continue
            prof_ok = True
            current = contrato.dtInicioContrato
            while current <= contrato.dtFimContrato:
                if current.weekday() == weekday:
                    aula = aulas_by_key.get((current, inicio, fim, prof_id))
                    if aula:
                        reservadas = models.Reserva.objects.filter(aulaSessao=aula, status="RESERVADA").count()
                        cap = aula.capacidade_efetiva()
                    else:
                        reservadas = 0
                        cap = _capacidade_para(horarios_list, prof_id)
                    if reservadas >= (cap or 0):
                        prof_ok = False
                        break
                current = current + timedelta(days=1)
            if prof_ok:
                slot_ok = True
                break
        if slot_ok:
            slots[key] = {
                "weekday": weekday,
                "inicio": inicio,
                "fim": fim,
                "allowed_profs": sorted(list(allowed)),
                "horarios": horarios_list,
            }

    if not slots and not horarios.exists():
        slots = {}
        for aula in aulas:
            reservadas = models.Reserva.objects.filter(aulaSessao=aula, status="RESERVADA").count()
            if reservadas >= aula.capacidade_efetiva():
                continue
            key = (aula.data.weekday(), aula.horaInicio, aula.horaFim)
            slots[key] = {"weekday": key[0], "inicio": key[1], "fim": key[2], "allowed_profs": [], "horarios": []}

    if request.method == "POST":
        escolhidas = []
        for idx in range(1, aulas_por_semana + 1):
            valor = request.POST.get(f"slot_{idx}") or ""
            if valor:
                escolhidas.append((idx, valor))
        if len(escolhidas) != aulas_por_semana:
            messages.error(request, f"Selecione exatamente {aulas_por_semana} horarios por semana.")
        else:
            conflitos = []
            faltando_prof = False
            usados = set()
            for idx, slot_value in escolhidas:
                if slot_value in usados:
                    messages.error(request, "Nao repita o mesmo horario.")
                    return redirect("contratos_agenda", pk=contrato.id)
                usados.add(slot_value)
                weekday, inicio, fim = slot_value.split("|")
                weekday = int(weekday)
                prof_id = request.POST.get(f"prof_for_{idx}") or ""
                try:
                    prof_id = int(prof_id)
                except ValueError:
                    prof_id = None
                if not prof_id:
                    faltando_prof = True
                    continue
                try:
                    inicio_time = datetime.strptime(inicio, "%H:%M").time()
                    fim_time = datetime.strptime(fim, "%H:%M").time()
                except ValueError:
                    messages.error(request, "Horario invalido na selecao.")
                    return redirect("contratos_agenda", pk=contrato.id)
                horario_key = (weekday, inicio_time, fim_time)
                slot_payload = slots.get(horario_key, {})
                allowed = set(slot_payload.get("allowed_profs") or [])
                horarios_list = slot_payload.get("horarios") or []
                if allowed and prof_id not in allowed:
                    messages.error(request, "Professor invalido para o horario selecionado.")
                    return redirect("contratos_agenda", pk=contrato.id)
                horario_config = next((h for h in horarios_list if h.profissional_id == prof_id), None) or next(
                    (h for h in horarios_list if h.profissional_id is None),
                    None,
                )
                current = contrato.dtInicioContrato
                while current <= contrato.dtFimContrato:
                    if current.weekday() == weekday:
                        aula = models.AulaSessao.objects.filter(
                            unidade=contrato.cdUnidade,
                            tipoServico=plano.cdTipoServico,
                            profissional_id=prof_id,
                            data=current,
                            horaInicio=inicio_time,
                            horaFim=fim_time,
                        ).first()
                        try:
                            if not aula:
                                aula = models.AulaSessao.objects.create(
                                    unidade=contrato.cdUnidade,
                                    tipoServico=plano.cdTipoServico,
                                    profissional_id=prof_id,
                                    data=current,
                                    horaInicio=inicio_time,
                                    horaFim=fim_time,
                                    capacidade=getattr(horario_config, "capacidade", None),
                                )
                            else:
                                update_fields = ["profissional"]
                                aula.profissional_id = prof_id
                                if aula.capacidade is None and horario_config and horario_config.capacidade is not None:
                                    aula.capacidade = horario_config.capacidade
                                    update_fields.append("capacidade")
                                aula.save(update_fields=update_fields)
                            if models.Reserva.objects.filter(aluno=contrato.cdAluno, aulaSessao=aula).exists():
                                current = current + timedelta(days=1)
                                continue
                            services.create_reserva(contrato.cdAluno, aula, status="RESERVADA")
                        except Exception:
                            conflitos.append(f"Sem vaga em {current} {inicio}")
                    current = current + timedelta(days=1)
            if faltando_prof:
                messages.error(request, "Selecione o professor em todos os horarios.")
                return redirect("contratos_agenda", pk=contrato.id)
            if conflitos:
                messages.warning(request, "Conflitos ao reservar: " + "; ".join(conflitos[:5]))
            else:
                messages.success(request, "Agenda criada com sucesso.")
                return redirect("alunos_detail", pk=contrato.cdAluno_id)

    weekday_labels = {
        0: "Segunda",
        1: "Terca",
        2: "Quarta",
        3: "Quinta",
        4: "Sexta",
        5: "Sabado",
        6: "Domingo",
    }
    slots_by_day = {key: [] for key in weekday_labels}
    for item in sorted(slots.values(), key=lambda x: (x["weekday"], x["inicio"])):
        slots_by_day[item["weekday"]].append(
            {
                "label": f'{item["inicio"].strftime("%H:%M")} - {item["fim"].strftime("%H:%M")}',
                "value": f'{item["weekday"]}|{item["inicio"].strftime("%H:%M")}|{item["fim"].strftime("%H:%M")}',
            }
        )

    slot_options = []
    for item in sorted(slots.values(), key=lambda x: (x["weekday"], x["inicio"])):
        slot_options.append(
            {
                "label": f'{weekday_labels[item["weekday"]]} {item["inicio"].strftime("%H:%M")} - {item["fim"].strftime("%H:%M")}',
                "value": f'{item["weekday"]}|{item["inicio"].strftime("%H:%M")}|{item["fim"].strftime("%H:%M")}',
                "allowed_profs": item.get("allowed_profs") or [],
            }
        )

    context = {
        "contrato": contrato,
        "aluno": contrato.cdAluno,
        "aulas_por_semana": aulas_por_semana,
        "slot_indices": list(range(1, aulas_por_semana + 1)),
        "slots_by_day": slots_by_day,
        "weekday_labels": weekday_labels,
        "slot_options": slot_options,
        "profissionais": profissionais,
        "horarios_configurados": horarios.exists(),
        "breadcrumbs": [("Home", reverse("dashboard")), ("Contratos", reverse("contratos_list")), ("Agenda", "#")],
        "active_menu": "cadastros",
    }
    return render(request, "contratos/agenda.html", context)


@login_required
def email_config_view(request):
    cfg = models.EmailConfiguracao.objects.order_by("-dtCadastro").first()
    if request.method == "POST":
        data = request.POST.copy()
        data = _inject_cd_value(models.EmailConfiguracao, data)
        form = forms.EmailConfiguracaoForm(data, instance=cfg)
        if form.is_valid():
            form.save()
            messages.success(request, "Configuracao de email salva.")
            return redirect("email_config")
        messages.error(request, "Verifique os erros do formulario.")
    else:
        form = forms.EmailConfiguracaoForm(instance=cfg)
    return render(
        request,
        "configuracoes/email.html",
        {
            "form": form,
            "title": "Configuracao de Email",
            "breadcrumbs": [("Home", reverse("dashboard")), ("Configuracoes", "#"), ("Email", "#")],
            "active_menu": "configuracoes",
        },
    )


def contrato_assinar(request, token):
    try:
        contrato_id = services.validar_token_contrato(token)
    except Exception:
        return render(request, "contratos/assinatura.html", {"token_invalido": True})
    contrato = get_object_or_404(models.Contrato, pk=contrato_id)
    if request.method == "POST":
        if contrato.status in ("ASSINADO", "ASSINADO_DIGITALMENTE"):
            return render(
                request,
                "contratos/assinatura_sucesso.html",
                {"contrato": contrato, "ja_assinado": True},
            )
        assinatura_nome = request.POST.get("assinatura_nome", "").strip()
        assinatura_documento = request.POST.get("assinatura_documento", "").strip()
        assinatura_data = request.POST.get("assinatura_data", "").strip()
        assinatura_ip = request.META.get("REMOTE_ADDR")
        if assinatura_nome and assinatura_data.startswith("data:image/"):
            try:
                from django.core.files.base import ContentFile
                import base64

                header, data = assinatura_data.split(",", 1)
                content = ContentFile(base64.b64decode(data), name=f"contrato_{contrato.id}.png")
                contrato.assinatura_imagem = content
            except Exception:
                assinatura_data = ""
        contrato.assinatura_nome = assinatura_nome
        contrato.assinatura_documento = assinatura_documento
        contrato.assinatura_ip = assinatura_ip
        contrato.status = "ASSINADO_DIGITALMENTE"
        contrato.assinado_em = timezone.now()
        contrato.save()
        return render(request, "contratos/assinatura_sucesso.html", {"contrato": contrato})
    html = services.render_contrato_html(contrato)
    return render(
        request,
        "contratos/assinatura.html",
        {"contrato": contrato, "contrato_html": html, "token": token},
    )


@login_required
def contrato_enviar_email(request, pk):
    contrato = get_object_or_404(models.Contrato, pk=pk)
    if request.method != "POST":
        return redirect("alunos_detail", pk=contrato.cdAluno_id)
    if not contrato.cdAluno.dsEmail:
        messages.warning(request, "Aluno sem email cadastrado.")
        return redirect("alunos_detail", pk=contrato.cdAluno_id)
    if services.enviar_contrato_para_assinatura(contrato, request.build_absolute_uri("/")):
        messages.success(request, "Contrato enviado para assinatura por email.")
    else:
        messages.warning(request, "Nao foi possivel enviar o email do contrato.")
    return redirect("alunos_detail", pk=contrato.cdAluno_id)


@login_required
def contrato_assinar_local(request, pk):
    contrato = get_object_or_404(models.Contrato, pk=pk)
    token = services.gerar_token_contrato(contrato)
    return redirect("contrato_assinar", token=token)


@login_required
def contrato_assinatura_detalhe(request, pk):
    contrato = get_object_or_404(models.Contrato, pk=pk)
    contrato_html = services.render_contrato_html(contrato)
    return render(
        request,
        "contratos/assinatura_detalhe.html",
        {
            "contrato": contrato,
            "contrato_html": contrato_html,
            "breadcrumbs": [("Home", reverse("dashboard")), ("Alunos", reverse("alunos_list")), ("Assinatura", "#")],
            "active_menu": "cadastros",
        },
    )
